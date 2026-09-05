import datetime
import io
import os

import openpyxl
import pytest

from tiaoke import output, samples
from tiaoke import timetable as tt_mod
from tiaoke.models import Event, Slot, SubLeg
from tiaoke.xlsx_writer import write_sheet

D = datetime.date

_REF = r"C:\Users\lolola\Desktop\調課代課程式\115-1手動調代課-兼課.xlsx"


def _render(event, timetable=None) -> openpyxl.worksheet.worksheet.Worksheet:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = event.sheet_name
    write_sheet(ws, event, timetable)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return openpyxl.load_workbook(buf).active


def _find(ws, text):
    return [c.coordinate for row in ws.iter_rows() for c in row
            if isinstance(c.value, str) and text in c.value]


def test_sheet_basic_shape():
    ws = _render(samples.get("瑞文1150223"))
    assert ws.title == "瑞文1150223"
    # 欄寬取自「範例1」原始值
    assert ws.column_dimensions["I"].width == pytest.approx(17.90625)
    assert ws.column_dimensions["A"].width == pytest.approx(12.26953125)
    assert ws.column_dimensions["E"].width == pytest.approx(13.0)
    assert ws.print_area in ("A1:I%d" % ws.max_row, "'瑞文1150223'!$A$1:$I$%d" % ws.max_row)


def test_banner_and_announce_text():
    ws = _render(samples.get("瑞文1150223"))
    assert _find(ws, "教  師  調  代  課  通  知  單 假單編號：手動  余瑞文 其他")
    assert _find(ws, "公告日期：115年")
    assert _find(ws, "02月13日")


def test_dates_stored_as_datetime_with_format():
    ws = _render(samples.get("瑞文1150223"))
    date_cells = [c for row in ws.iter_rows() for c in row
                  if c.number_format == 'm"月"d"日"']
    assert date_cells
    for c in date_cells:
        assert hasattr(c.value, "year")  # 真正的日期物件，不是字串


def test_note_row_present_only_when_note_set():
    with_note = _render(samples.get("若耶1150226"))
    assert _find(with_note, "說明:")
    without = _render(samples.get("瑞文1150223"))
    assert not _find(without, "說明:")


def test_class_slip_footer():
    ws = _render(samples.get("瑞文1150223"))
    assert _find(ws, "* 請學藝股長公佈。")


def test_merged_regions_exist():
    ws = _render(samples.get("瑞文1150223"))
    merged = {str(r) for r in ws.merged_cells.ranges}
    assert "B1:I1" in merged           # 教師單橫幅
    assert any(m.startswith("E2:E3") for m in merged)  # 授課科目
    # 「調課後授課時間」「原授課時間」不合併（比照範例檔）
    assert "C2:D2" not in merged
    assert "G2:H2" not in merged


@pytest.mark.skipif(not os.path.exists(_REF), reason="找不到範例檔")
def test_example1_matches_reference_layout():
    ws = _render(samples.get("範例1"))
    ex = openpyxl.load_workbook(_REF)["範例1"]

    def merges(w):
        return sorted(str(r) for r in w.merged_cells.ranges
                      if not str(r).startswith(("K", "L")))

    assert merges(ws) == merges(ex)
    assert ws.max_row == 19
    for col in "ABCDEFGHI":
        assert ws.column_dimensions[col].width == pytest.approx(
            ex.column_dimensions[col].width)


def test_write_to_master_creates_then_replaces(tmp_path):
    master = tmp_path / "master.xlsx"
    ev = samples.get("瑞文1150223")

    r1 = output.write_to_master(ev, str(master))
    assert r1.ok and not r1.replaced_sheet
    assert master.exists()

    r2 = output.write_to_master(ev, str(master))
    assert r2.ok and r2.replaced_sheet

    wb = openpyxl.load_workbook(master)
    assert wb.sheetnames.count("瑞文1150223") == 1


def test_run_requires_a_target():
    with pytest.raises(ValueError):
        output.run(samples.get("瑞文1150223"))


def test_run_both_targets(tmp_path):
    master = tmp_path / "m.xlsx"
    new = tmp_path / "n.xlsx"
    results = output.run(
        samples.get("代課範例"),
        to_master=True, master_path=str(master),
        save_new=True, dest_path=str(new),
    )
    assert len(results) == 2
    assert all(r.ok for r in results)
    assert master.exists() and new.exists()


def _timetable_with_note(weekday: int, period: int, note: str, teacher="余瑞文"):
    table = tt_mod.Timetable()
    table.teachers[teacher] = tt_mod.TeacherTable(name=teacher, slots=[
        tt_mod.Slot(weekday=weekday, period=period, subject="健康與護理",
                    klass="電機一", note=note),
    ])
    return table


def test_j_column_marks_extra_hours_blue_bold_kaiu():
    # 2026-09-07 是星期一
    ev = Event("余瑞文", "病假", "手動", D(2026, 9, 3), legs=[
        SubLeg("電機一", "余瑞文", "健康與護理", Slot(D(2026, 9, 7), 1), "王小明"),
    ])
    ws = _render(ev, timetable=_timetable_with_note(1, 1, "(兼)"))
    j_cells = [c for c in ws["J"] if c.value not in (None, "")]
    # 原老師、代課老師、班級單三張都要有
    assert [c.value for c in j_cells] == ["兼課", "兼課", "兼課"]
    for c in j_cells:
        assert c.font.bold is True
        assert c.font.name == "標楷體"
        assert c.font.color.rgb.endswith("0000FF")


def test_j_column_empty_without_timetable():
    ev = Event("余瑞文", "病假", "手動", D(2026, 9, 3), legs=[
        SubLeg("電機一", "余瑞文", "健康與護理", Slot(D(2026, 9, 7), 1), "王小明"),
    ])
    ws = _render(ev)  # 沒給課表
    j_values = [c.value for c in ws["J"] if c.value not in (None, "")]
    assert j_values == []


def test_j_column_not_in_print_area():
    ev = Event("余瑞文", "病假", "手動", D(2026, 9, 3), legs=[
        SubLeg("電機一", "余瑞文", "健康與護理", Slot(D(2026, 9, 7), 1), "王小明"),
    ])
    ws = _render(ev, timetable=_timetable_with_note(1, 1, "(兼)"))
    assert ws.print_area in ("A1:I%d" % ws.max_row,
                             "'%s'!$A$1:$I$%d" % (ws.title, ws.max_row))
