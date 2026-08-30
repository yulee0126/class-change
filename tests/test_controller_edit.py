import datetime

import openpyxl

from tiaoke import samples
from tiaoke.models import SwapLeg
from tiaoke.ui.controller import AppController

D = datetime.date


def _ctl_with_swap():
    c = AppController()
    c.new_event()
    c.update_event_fields(originator="余瑞文", form_no="手動")
    c.add_swap_leg(
        klass="高一甲", teacher_a="余瑞文", subject_a="健護",
        date_a=D(2026, 2, 23), period_a=1,
        teacher_b="洪瑞霞", subject_b="班會",
        date_b=D(2026, 2, 25), period_b=5,
    )
    return c


def test_leg_form_data_round_trips_into_update():
    c = _ctl_with_swap()
    data = c.leg_form_data(0)
    assert data["kind"] == "swap"
    assert data["teacher_a"] == "余瑞文"
    assert data["period_b"] == 5

    # 改乙節次 5 → 6
    data.pop("kind")
    data["period_b"] = 6
    c.update_leg(0, "swap", **data)

    leg = c.current.legs[0]
    assert isinstance(leg, SwapLeg)
    assert leg.slot_b.period == 6
    assert len(c.current.legs) == 1  # 是取代，不是新增


def test_update_leg_can_change_type():
    c = _ctl_with_swap()
    c.update_leg(0, "sub", klass="高一甲", orig_teacher="余瑞文", subject="健護",
                 date=D(2026, 2, 23), period=1, sub_teacher="王小明")
    from tiaoke.models import SubLeg
    assert isinstance(c.current.legs[0], SubLeg)


def test_export_all_xlsx(tmp_path):
    c = AppController()
    c.project.events = [samples.get("瑞文1150223"), samples.get("代課範例"),
                        samples.get("炆明1150831")]
    out = tmp_path / "全部.xlsx"
    r = c.export_all_xlsx(str(out))
    assert r.ok
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["瑞文1150223", "建勳1110119", "炆明1150831"]


def test_export_all_dedupes_sheet_names(tmp_path):
    c = AppController()
    c.project.events = [samples.get("瑞文1150223"), samples.get("瑞文1150223")]
    out = tmp_path / "dup.xlsx"
    r = c.export_all_xlsx(str(out))
    assert r.ok
    wb = openpyxl.load_workbook(out)
    assert len(wb.sheetnames) == 2 and len(set(wb.sheetnames)) == 2
