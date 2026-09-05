"""把 xlsx_writer 產生的通知單 Excel 讀回 Event，供搜尋既有檔案後編輯／產製報表用。

設計原則：**完全不解析「備註」欄的文字**。備註是自由文字，使用者常會手改
（例如附註「(蓁妍同步調)」、把兩個人的名字用「/」接在一起），這些編輯不影響
還原的正確性 —— 因為每一列的「誰、班級、日期、節次、是否反白」全部是結構化
資料（欄位值／儲存格底色），不是文字。備註欄只在還原「說明」整段時會用到
（`說明:\n...` 那一列），單列的備註文字則整個略過不讀。

還原邏輯（只看教師調代課通知單，班級單是衍生資料，略過不讀）：
  · 一列「調課後時間、原時間都有值」＝對調事件的一邊
  · 一列「只有原時間」＝代課腳的「被代課」那邊
  · 一列「只有新時間」＝代課腳的「代課」那邊
  · 同一 (班級, {new,orig}) 這兩個時段是同一次調課事件——依 (new,orig) 的方向
    分成 A、B 兩側，每側各自的人數才是重點：
      - 兩側都恰好 1 人 → 一般 SwapLeg
      - 任一側 2 人以上 → 協同教學（P7）：那幾位老師各自登記一列一模一樣的資料
        （同 klass/new/orig，只差老師名字），對方教師單卻只有一列。若呼叫端有
        給 `timetable`（已跑過教師配當表比對）且能用 `Slot.co_teachers` 確認
        「多人那側彼此互為協同老師」，就整組合併成一個 CoSwapLeg；驗不到協同
        關係就不合併，整組列為 unmatched
  · 一個「被代課」與一個「代課」在同一 (klass, slot) → 配成一個 SubLeg
    （from_swap 由儲存格底色是否反白決定，不是文字）
如果配不出（例如同一節被兩位老師同時頂替、且查不到協同關係之類，超出目前
資料模型能表示的範圍），視為整份解析失敗並丟出 ParseError，不回傳猜測、可能
錯誤的資料。
"""

from __future__ import annotations

import datetime
import os
import re

from openpyxl import load_workbook

from . import roc
from .models import CoSwapLeg, Event, Slot, SubLeg, SwapLeg
from .styles import CLASS_TITLE_TEXT

_BANNER_RE = re.compile(r"假單編號：(?P<form_no>.+?)  (?P<originator>.+) (?P<leave_type>\S+)$")
_ANNOUNCE_YEAR_RE = re.compile(r"公告日期：(\d+)年")
_ANNOUNCE_MD_RE = re.compile(r"(\d+)月(\d+)日")


class ParseError(Exception):
    """檔案格式看不懂，或有資料配不出結構化的調課／代課關係。"""


# --------------------------------------------------------------------------

def read_event(path: str, sheet_name: str | None = None, timetable=None) -> Event:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
    fname = os.path.basename(path)

    entries, meta, note_text = _scan(ws)
    if not meta:
        raise ParseError(f"{fname}：找不到教師調代課通知單的版面，無法解析。")

    legs, unmatched = _pair_entries(entries, timetable)
    if unmatched:
        detail = "、".join(_describe(e) for e in unmatched)
        raise ParseError(
            f"{fname}：有 {len(unmatched)} 列配不出調課／代課關係（可能是超出目前"
            f"資料模型的特殊多人調課），無法完整還原：{detail}")

    return Event(
        originator=meta.get("originator", ""),
        leave_type=meta.get("leave_type", ""),
        form_no=meta.get("form_no", ""),
        announce_date=meta.get("announce_date") or datetime.date.today(),
        note=note_text,
        class_slip_style=_detect_class_style(ws),
        sheet_name_override=ws.title,
        legs=legs,
    )


def read_events(path: str, timetable=None) -> list[tuple[str, Event | None, str]]:
    """讀一個活頁簿裡所有分頁。回傳 [(分頁名, Event 或 None, 錯誤訊息或"")]。

    單一分頁解析失敗不影響其他分頁（供「產製報表」掃資料夾用）。
    """
    wb = load_workbook(path, data_only=True)
    out: list[tuple[str, Event | None, str]] = []
    for ws in wb.worksheets:
        try:
            out.append((ws.title, read_event(path, sheet_name=ws.title, timetable=timetable), ""))
        except ParseError as exc:
            out.append((ws.title, None, str(exc)))
    return out


def _describe(e: dict) -> str:
    slot = e["new"] or e["orig"]
    where = roc.slot_label(slot.date, slot.period) if slot else "?"
    return f"{e['teacher']}/{e['klass']}/{where}"


# --------------------------------------------------------------------------
# 掃描教師調代課通知單，抓出每一列的結構化資料
# --------------------------------------------------------------------------

def _scan(ws) -> tuple[list[dict], dict, str]:
    entries: list[dict] = []
    meta: dict = {}
    note_text = ""
    r, max_r = 1, ws.max_row

    while r <= max_r:
        a = ws.cell(r, 1).value
        c_next = ws.cell(r + 1, 3).value if r + 1 <= max_r else None
        if not (a and c_next == "調課後授課時間"):
            r += 1
            continue

        teacher = str(a).strip()
        if "form_no" not in meta:
            meta.update(_parse_banner(str(ws.cell(r, 2).value or "")))
        r += 3  # 跳過標題列 + 兩列表頭

        while r <= max_r:
            h_val = ws.cell(r, 8).value
            if isinstance(h_val, str) and h_val.startswith("公告日期"):
                if "announce_date" not in meta:
                    meta["announce_date"] = _parse_announce(h_val, ws.cell(r, 9).value)
                r += 1
                nxt = ws.cell(r, 1).value if r <= max_r else None
                if isinstance(nxt, str) and nxt.startswith("說明:"):
                    if not note_text:
                        note_text = nxt[len("說明:"):].lstrip("\n")
                    r += 1
                break

            a_val = ws.cell(r, 1).value
            if a_val in (None, ""):
                break
            entries.append(_read_data_row(ws, r, str(a_val).strip(), teacher))
            r += 1

    return entries, meta, note_text


def _read_data_row(ws, r: int, klass: str, teacher: str) -> dict:
    b_val = ws.cell(r, 2).value
    new_slot = Slot(_to_date(b_val), int(ws.cell(r, 4).value)) if b_val not in (None, "") else None
    subject = str(ws.cell(r, 5).value or "").strip()
    f_val = ws.cell(r, 6).value
    orig_slot = Slot(_to_date(f_val), int(ws.cell(r, 8).value)) if f_val not in (None, "") else None
    fill = ws.cell(r, 1).fill
    highlight = bool(fill and fill.patternType == "solid")
    return dict(teacher=teacher, klass=klass, subject=subject,
               new=new_slot, orig=orig_slot, highlight=highlight)


def _to_date(v) -> datetime.date:
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    raise ParseError(f"日期格式看不懂：{v!r}")


def _parse_banner(banner: str) -> dict:
    m = _BANNER_RE.search(banner)
    if not m:
        raise ParseError(f"橫幅看不懂，抓不到假單編號／發起教師／假別：{banner!r}")
    return {"form_no": m.group("form_no"), "originator": m.group("originator"),
           "leave_type": m.group("leave_type")}


def _parse_announce(h_val, i_val) -> datetime.date:
    my = _ANNOUNCE_YEAR_RE.search(str(h_val or ""))
    md = _ANNOUNCE_MD_RE.search(str(i_val or ""))
    if not (my and md):
        raise ParseError(f"公告日期看不懂：{h_val!r} {i_val!r}")
    return datetime.date(int(my.group(1)) + 1911, int(md.group(1)), int(md.group(2)))


def _detect_class_style(ws) -> str:
    marker = CLASS_TITLE_TEXT.strip()
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and marker in c.value:
                return "title"
    return "banner"


# --------------------------------------------------------------------------
# 把「調課後時間／原時間」列配對回 SwapLeg／SubLeg
# --------------------------------------------------------------------------

def _pair_entries(entries: list[dict], timetable=None) -> tuple[list, list[dict]]:
    swap_side = [e for e in entries if e["new"] and e["orig"]]
    sub_out = [e for e in entries if e["orig"] and not e["new"]]     # 被代課
    sub_in = [e for e in entries if e["new"] and not e["orig"]]      # 代課者
    orphans = [e for e in entries if not e["new"] and not e["orig"]]  # 兩邊都空，異常

    legs: list = []
    unmatched: list[dict] = list(orphans)

    # 同一 (班級, {new,orig}) 這兩個時段就是同一次調課事件；一側只有一人＝一般
    # SwapLeg，任一側多人就是協同教學（P7）——這種情況下協同的那幾位老師各自
    # 登記一列一模一樣的資料（同 klass/new/orig，只差老師名字），對方教師單卻
    # 只有一列，所以「側」要用整組人數判斷，不是逐列貪婪配對。
    groups: dict[tuple, list[dict]] = {}
    for e in swap_side:
        key = (e["klass"], frozenset([e["new"], e["orig"]]))
        groups.setdefault(key, []).append(e)

    for (klass, _pair), group in groups.items():
        x, y = group[0]["new"], group[0]["orig"]
        side_a = [e for e in group if e["new"] == x and e["orig"] == y]
        side_b = [e for e in group if e["new"] == y and e["orig"] == x]

        if len(side_a) == 1 and len(side_b) == 1:
            ea, eb = side_a[0], side_b[0]
            legs.append(SwapLeg(
                klass=klass,
                teacher_a=ea["teacher"], subject_a=ea["subject"], slot_a=ea["orig"],
                teacher_b=eb["teacher"], subject_b=eb["subject"], slot_b=eb["orig"],
            ))
        elif side_a and side_b and _all_mutual_co_teachers(timetable, klass, side_a) \
                and _all_mutual_co_teachers(timetable, klass, side_b):
            legs.append(CoSwapLeg(
                klass=klass,
                teachers_a=[e["teacher"] for e in side_a], subject_a=side_a[0]["subject"],
                slot_a=side_a[0]["orig"],
                teachers_b=[e["teacher"] for e in side_b], subject_b=side_b[0]["subject"],
                slot_b=side_b[0]["orig"],
            ))
        else:
            unmatched.extend(group)

    used_in = [False] * len(sub_in)
    for e_out in sub_out:
        matched = False
        for j, e_in in enumerate(sub_in):
            if used_in[j]:
                continue
            if e_in["klass"] == e_out["klass"] and e_in["new"] == e_out["orig"]:
                legs.append(SubLeg(
                    klass=e_out["klass"], orig_teacher=e_out["teacher"],
                    subject=e_out["subject"], slot=e_out["orig"],
                    sub_teacher=e_in["teacher"],
                    from_swap=bool(e_out["highlight"] or e_in["highlight"]),
                ))
                used_in[j] = True
                matched = True
                break
        if not matched:
            unmatched.append(e_out)
    unmatched += [e for j, e in enumerate(sub_in) if not used_in[j]]

    return legs, unmatched


def _all_mutual_co_teachers(timetable, klass: str, side_entries: list[dict]) -> bool:
    """side_entries 只有 1 人：不用驗證，直接算數（一般情況）。
    2 人以上：要求彼此兩兩都能在課表／配當表比對中確認互為協同老師，才允許
    合併成一個 CoSwapLeg——只要有一對驗不到，就不亂猜，交給呼叫端當unmatched。"""
    if len(side_entries) < 2:
        return True
    if timetable is None:
        return False
    names = [e["teacher"] for e in side_entries]
    for i, a in enumerate(names):
        for b in names[i + 1:]:
            if not (_co_teaches(timetable, a, b, klass) or _co_teaches(timetable, b, a, klass)):
                return False
    return True


def _co_teaches(timetable, teacher: str, other: str, klass: str) -> bool:
    """teacher／other 是否曾在同一班比對出協同關係（不強求跟這次異動同星期節次——
    調課常常永久改掉原本的星期節次，比對時課表可能已經不再有那個舊時段了，
    只要兩人在這個班有登記過協同，就採信）。"""
    t = timetable.teachers.get(teacher)
    if not t:
        return False
    return any(s.klass == klass and other in s.co_teachers for s in t.slots)
