import datetime
import os

import openpyxl
import pytest

from tiaoke import output, samples
from tiaoke.models import CoSwapLeg, Slot, SubLeg, SwapLeg
from tiaoke.ui.controller import AppController
from tiaoke.xlsx_reader import ParseError, read_event, read_events
from tiaoke.xlsx_writer import write_sheet

D = datetime.date

# 這台機器上的實體課表 PDF／教師配當表（若不在則跳過相關測試）
_PDF = r"C:\Users\lolola\Desktop\1151教師課表_正式公布.pdf"
_HAS_PDF = os.path.exists(_PDF)
_PEIDANG = r"C:\Users\lolola\Desktop\115-1教師配當表簽稿.xlsx"
_HAS_PEIDANG = os.path.exists(_PEIDANG)


def _write(event, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = event.sheet_name
    write_sheet(ws, event)
    wb.save(path)


def _leg_set(legs):
    """轉成可比較、不受順序影響的集合。"""
    out = set()
    for leg in legs:
        if isinstance(leg, SwapLeg):
            # teacher_a/teacher_b 的順序在還原時不保證跟原本一樣，兩種都收進來比對其一即可
            out.add(("swap", leg.klass, frozenset([
                (leg.teacher_a, leg.subject_a, leg.slot_a),
                (leg.teacher_b, leg.subject_b, leg.slot_b),
            ])))
        else:
            out.add(("sub", leg.klass, leg.orig_teacher, leg.subject, leg.slot,
                     leg.sub_teacher, leg.from_swap))
    return out


@pytest.mark.parametrize("name", list(samples.SAMPLES))
def test_round_trip_all_samples(name, tmp_path):
    ev = samples.get(name)
    path = tmp_path / "x.xlsx"
    _write(ev, str(path))

    back = read_event(str(path))
    assert back.originator == ev.originator
    assert back.leave_type == ev.leave_type
    assert back.form_no == ev.form_no
    assert back.announce_date == ev.announce_date
    assert back.class_slip_style == ev.class_slip_style
    assert back.sheet_name_override == ev.sheet_name
    assert _leg_set(back.legs) == _leg_set(ev.legs)


def test_round_trip_preserves_note(tmp_path):
    ev = samples.get("若耶1150226")
    assert ev.note
    path = tmp_path / "x.xlsx"
    _write(ev, str(path))
    back = read_event(str(path))
    assert back.note == ev.note


def test_hand_edited_note_text_does_not_break_parsing(tmp_path):
    """使用者手改教師單備註欄（附註、合併姓名）不影響結構化還原。"""
    ev = samples.get("炆明1150831")
    path = tmp_path / "x.xlsx"
    _write(ev, str(path))

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    edited = 0
    for row in ws.iter_rows():
        cell = row[8]  # I 欄＝備註
        if isinstance(cell.value, str) and cell.value.endswith("調課"):
            cell.value = cell.value + "\n(蓁妍同步調)"
            edited += 1
    assert edited > 0
    wb.save(path)

    back = read_event(str(path))
    assert _leg_set(back.legs) == _leg_set(ev.legs)


def test_from_swap_flag_recovered_from_highlight_fill(tmp_path):
    ev = samples.get("炆明1150831")
    hl_subs = [l for l in ev.legs if isinstance(l, SubLeg) and l.from_swap]
    assert hl_subs  # 樣本本身要有這種列，測試才有意義

    path = tmp_path / "x.xlsx"
    _write(ev, str(path))
    back = read_event(str(path))
    recovered = [l for l in back.legs if isinstance(l, SubLeg) and l.from_swap]
    assert len(recovered) == len(hl_subs)


def test_read_events_multi_sheet(tmp_path):
    events = [samples.get("瑞文1150223"), samples.get("代課範例")]
    path = tmp_path / "all.xlsx"
    output.export_all(events, str(path))

    results = read_events(str(path))
    assert len(results) == 2
    names = {name for name, ev, err in results}
    assert names == {ev.sheet_name for ev in events}
    assert all(err == "" and ev is not None for _n, ev, err in results)


def test_unmatched_swap_side_raises_parse_error(tmp_path):
    """故意弄壞一份對調單（刪掉對方那張教師單），該有配不出來的錯誤。"""
    ev = samples.get("瑞文1150223")
    path = tmp_path / "x.xlsx"
    _write(ev, str(path))

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    del wb[ws.title]
    ws2 = wb.create_sheet(ws.title)
    # 只搬「余瑞文」那張教師單（第一張），把其餘對方老師的單跟班級單都丟掉
    for row in ws.iter_rows(min_row=1, max_row=6):
        for cell in row:
            ws2.cell(cell.row, cell.column, cell.value)
            if cell.fill and cell.fill.patternType:
                ws2.cell(cell.row, cell.column).fill = cell.fill
    wb.save(path)

    with pytest.raises(ParseError, match="配不出"):
        read_event(str(path))


# ==========================================================================
# 協同教學（P7 F4）：兩位協同老師各自登記一列跟同一位目標老師對調
# ==========================================================================

def _co_swap_event():
    """模擬 _CoSwapForm／add_co_swap 產生的樣子：趙瑋、周蓁妍協同，一起跟張宥恩對調。"""
    from tiaoke.models import Event, Slot
    return Event(
        originator="趙瑋", leave_type="病假", form_no="手動", announce_date=D(2026, 9, 1),
        legs=[
            SwapLeg("綜職二", "趙瑋", "基礎雜糧加工實作", Slot(D(2026, 9, 3), 5),
                   "張宥恩", "物品整理實務", Slot(D(2026, 9, 1), 5)),
            SwapLeg("綜職二", "周蓁妍", "基礎雜糧加工實作", Slot(D(2026, 9, 3), 5),
                   "張宥恩", "物品整理實務", Slot(D(2026, 9, 1), 5)),
        ],
    )


def _co_teach_timetable():
    from tiaoke.timetable import Slot as TTSlot, TeacherTable, Timetable
    table = Timetable()
    table.teachers["趙瑋"] = TeacherTable(name="趙瑋", slots=[
        TTSlot(4, 5, "基礎雜糧加工實作", "綜職二", co_teachers=["周蓁妍"]),
    ])
    table.teachers["周蓁妍"] = TeacherTable(name="周蓁妍", slots=[
        TTSlot(4, 5, "基礎雜糧加工實作", "綜職二", co_teachers=["趙瑋"]),
    ])
    return table


def _dedupe_target_row(path) -> None:
    """`add_co_swap` 產生的檔案裡，張宥恩（目標老師）本來會有 2 列一模一樣的資料
    （趙瑋、周蓁妍各自對調一筆都指向他）。真實的歷史檔案是人工把這種重複列
    刪成只留一列（另一邊改用「與趙瑋/蓁妍老師調課」這種手改備註合併表示）——
    這裡把其中一列清空，重現那個「兩邊搶同一列」的真實情境。

    用清空儲存格而不是 ws.delete_rows()：openpyxl 的 delete_rows 不會妥善調整
    其他地方的 merged_cells 範圍，會把後面分頁的合併儲存格弄亂，連帶讓 _scan
    讀不到後面的教師單。
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    hits = [row[0].row for row in ws.iter_rows(min_row=1)
           if row[0].value == "綜職二" and row[4].value == "物品整理實務"]
    assert len(hits) == 2, f"預期張宥恩有 2 列重複資料，實際找到 {hits}"
    for col in range(1, 10):
        ws.cell(hits[1], col).value = None
    wb.save(path)


def test_co_swap_without_timetable_raises_parse_error(tmp_path):
    """沒有課表可查協同關係時，維持原本行為：配不出來就報錯，不亂猜。"""
    ev = _co_swap_event()
    path = tmp_path / "x.xlsx"
    _write(ev, str(path))
    _dedupe_target_row(path)
    with pytest.raises(ParseError, match="配不出"):
        read_event(str(path))


def test_co_swap_with_timetable_resolves_to_one_co_swap_leg(tmp_path):
    """有課表可查協同關係時，兩位協同老師合併還原成一個 CoSwapLeg（不是兩個 SwapLeg）。"""
    ev = _co_swap_event()
    path = tmp_path / "x.xlsx"
    _write(ev, str(path))
    _dedupe_target_row(path)

    back = read_event(str(path), timetable=_co_teach_timetable())
    assert len(back.legs) == 1
    leg = back.legs[0]
    assert isinstance(leg, CoSwapLeg)
    assert set(leg.teachers_a) == {"趙瑋", "周蓁妍"}
    assert leg.teachers_b == ["張宥恩"]
    assert leg.subject_a == "基礎雜糧加工實作"
    assert leg.subject_b == "物品整理實務"
    assert leg.slot_a == Slot(D(2026, 9, 3), 5)
    assert leg.slot_b == Slot(D(2026, 9, 1), 5)


def test_co_swap_unconfirmed_co_teach_stays_unmatched(tmp_path):
    """給了課表但驗不到協同關係時，整組（含目標老師那列）都算配不出，不會亂猜。"""
    ev = _co_swap_event()
    path = tmp_path / "x.xlsx"
    _write(ev, str(path))
    _dedupe_target_row(path)

    from tiaoke.timetable import TeacherTable, Timetable
    table = Timetable()
    table.teachers["趙瑋"] = TeacherTable(name="趙瑋")  # 沒有課表資料、查不到協同
    with pytest.raises(ParseError, match="配不出"):
        read_event(str(path), timetable=table)


@pytest.mark.skipif(not (_HAS_PEIDANG and _HAS_PDF), reason="找不到實體配當表或課表 PDF")
def test_real_co_teach_file_improves_with_timetable():
    """P6 留下的 1002-趙瑋1150903.xlsx 解析失敗案例，接上協同課表後應該改善（但不會全解開）。

    這份真實檔案裡趙瑋、周蓁妍協同教「基礎雜糧加工實作」（綜職二）的 3 節，
    配當表＋課表能確認協同關係，這 3 節現在正確合併成 3 個 CoSwapLeg 解開了。

    同一份檔案裡還有「數學」（綜職一）：趙瑋、周蓁妍、林冠佑三人也用同樣的
    手法各登記一列同步調課，但配當表完全沒提到這堂課是協同課，驗不到協同關係。
    這裡刻意不採用「先到先配、按貪婪順序把林冠佑配給趙瑋」這種舊做法——那樣
    等於隨機武斷地認定趙瑋才是真正跟林冠佑對調的人、悄悄丟掉周蓁妍那筆同樣
    的宣告，是更嚴重的錯誤。改成分組比對後，這 3 列（趙瑋、周蓁妍、林冠佑）
    會一起被列為配不出，而不是靜悄悄地猜一個可能是錯的答案。
    """
    from tiaoke import timetable as tt_mod

    path = r"C:\Users\lolola\Documents\GitHub\class-change\build_out\調代課單\1002-趙瑋1150903.xlsx"
    if not os.path.exists(path):
        pytest.skip("找不到真實範例檔 1002-趙瑋1150903.xlsx")

    table = tt_mod.parse_pdf(_PDF)
    rows = tt_mod.parse_co_teaching_xlsx(_PEIDANG)
    tt_mod.apply_co_teaching(table, rows)

    with pytest.raises(ParseError) as exc_info:
        read_event(path, timetable=table)
    msg = str(exc_info.value)
    assert "有 3 列配不出" in msg
    assert "趙瑋/綜職一/9/3(四)第1節" in msg
    assert "周蓁妍/綜職一/9/3(四)第1節" in msg
    assert "林冠佑/綜職一/9/3(四)第3節" in msg
    assert "綜職二" not in msg  # 基礎雜糧加工實作那 3 節（綜職二）現在解開了


@pytest.mark.skipif(not (_HAS_PEIDANG and _HAS_PDF), reason="找不到實體配當表或課表 PDF")
def test_real_co_teach_file_resolves_co_swap_legs():
    """確認解得開的那 3 節確實還原成 CoSwapLeg（teachers_a 含趙瑋、周蓁妍兩人）。"""
    from tiaoke import timetable as tt_mod

    path = r"C:\Users\lolola\Documents\GitHub\class-change\build_out\調代課單\1002-趙瑋1150903.xlsx"
    if not os.path.exists(path):
        pytest.skip("找不到真實範例檔 1002-趙瑋1150903.xlsx")

    table = tt_mod.parse_pdf(_PDF)
    rows = tt_mod.parse_co_teaching_xlsx(_PEIDANG)
    tt_mod.apply_co_teaching(table, rows)

    entries, meta, _note = _scan_real_sheet(path)
    legs, unmatched = _pair_entries_for_test(entries, table)
    co_legs = [l for l in legs if isinstance(l, CoSwapLeg) and l.klass == "綜職二"]
    assert len(co_legs) == 3
    for leg in co_legs:
        assert set(leg.teachers_a) == {"趙瑋", "周蓁妍"}
        assert leg.teachers_b == ["張宥恩"] or leg.teachers_b == ["張秉正"]


def _scan_real_sheet(path):
    from tiaoke.xlsx_reader import _scan
    wb = openpyxl.load_workbook(path, data_only=True)
    return _scan(wb.worksheets[0])


def _pair_entries_for_test(entries, timetable):
    from tiaoke.xlsx_reader import _pair_entries
    return _pair_entries(entries, timetable)


def test_missing_banner_raises_parse_error(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "不是通知單"
    path = tmp_path / "x.xlsx"
    wb.save(path)
    with pytest.raises(ParseError):
        read_event(str(path))


# ==========================================================================
# controller：搜尋既有檔案 / 讀回編輯 / 存回 / 產製報表
# ==========================================================================

def test_list_slip_files_sorted_newest_first_and_skips_lock_files(tmp_path):
    c = AppController()
    _write(samples.get("代課範例"), str(tmp_path / "a.xlsx"))
    os.utime(tmp_path / "a.xlsx", (1000, 1000))
    _write(samples.get("瑞文1150223"), str(tmp_path / "b.xlsx"))
    os.utime(tmp_path / "b.xlsx", (2000, 2000))
    (tmp_path / "~$a.xlsx").write_text("lock")
    (tmp_path / "notes.txt").write_text("x")

    names = c.list_slip_files(str(tmp_path))
    assert names == ["b.xlsx", "a.xlsx"]
    assert c.list_slip_files(str(tmp_path / "不存在")) == []
    assert c.list_slip_files("") == []


def test_load_event_from_file_adds_and_selects_event(tmp_path):
    ev = samples.get("代課範例")
    path = tmp_path / "x.xlsx"
    _write(ev, str(path))

    c = AppController()
    c.new_event()  # 原本已經有一張在編輯
    loaded = c.load_event_from_file(str(path))
    assert c.current is loaded
    assert loaded.originator == ev.originator
    assert getattr(loaded, "_source_path") == str(path)


def test_resave_loaded_file_same_path_overwrites(tmp_path):
    ev = samples.get("代課範例")
    path = tmp_path / "x.xlsx"
    _write(ev, str(path))

    c = AppController()
    c.load_event_from_file(str(path))
    result = c.resave_loaded_file(str(path))
    assert result.ok
    assert os.path.exists(path)
    assert getattr(c.current, "_source_path") == str(path)


def test_resave_loaded_file_new_name_deletes_old(tmp_path):
    ev = samples.get("代課範例")
    old_path = tmp_path / "old.xlsx"
    new_path = tmp_path / "new.xlsx"
    _write(ev, str(old_path))

    c = AppController()
    c.load_event_from_file(str(old_path))
    result = c.resave_loaded_file(str(new_path))
    assert result.ok
    assert not os.path.exists(old_path)
    assert os.path.exists(new_path)
    assert getattr(c.current, "_source_path") == str(new_path)


def test_generate_report_scans_folder_and_skips_marked_files(tmp_path):
    folder = tmp_path / "調代課單"
    folder.mkdir()
    _write(samples.get("代課範例"), str(folder / "1001-x.xlsx"))
    _write(samples.get("瑞文1150223"), str(folder / "1002-y.xlsx"))
    # 作廢檔（~~ 前綴）跟 Excel 鎖檔（~$ 前綴）都該略過
    _write(samples.get("炆明1150831"), str(folder / "~~1003-作廢.xlsx"))
    (folder / "~$1001-x.xlsx").write_text("lock")
    (folder / "說明.txt").write_text("not excel")

    out_folder = tmp_path / "record"
    c = AppController()
    rep = c.generate_report(str(folder), str(out_folder))

    assert rep.ok
    assert rep.files_ok == 2          # 只有 1001、1002 被算進去
    assert not rep.files_failed
    assert os.path.exists(rep.path)
    assert os.path.dirname(rep.path) == str(out_folder)
    assert "調代課記錄-" in os.path.basename(rep.path)

    wb = openpyxl.load_workbook(rep.path)
    detail = [r for r in wb["調代課明細"].iter_rows(min_row=2, values_only=True) if r[0]]
    assert len(detail) == rep.rows
    assert rep.rows == 1 + 6  # 代課範例 1 列（代課）+ 瑞文1150223 三筆對調各 2 列 = 6 列


def test_generate_report_reports_unparseable_files_without_crashing(tmp_path):
    folder = tmp_path / "調代課單"
    folder.mkdir()
    _write(samples.get("代課範例"), str(folder / "good.xlsx"))
    (folder / "bad.xlsx").write_bytes(b"not a real xlsx file")

    c = AppController()
    rep = c.generate_report(str(folder), str(tmp_path / "record"))
    assert rep.ok
    assert rep.files_ok == 1
    assert any("bad.xlsx" in f for f in rep.files_failed)


def test_generate_report_no_folder():
    c = AppController()
    rep = c.generate_report("", "")
    assert not rep.ok
    assert "資料夾" in rep.error
