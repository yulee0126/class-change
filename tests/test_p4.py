"""P4：先調後代（from_swap）→ 反白列（J 欄簡稱已取消輸出）。"""

import datetime
import io

import openpyxl

from tiaoke import samples
from tiaoke.builder import ClassSlip, TeacherSlip, build
from tiaoke.models import Event, Slot, SubLeg
from tiaoke.xlsx_writer import write_sheet

D = datetime.date


def _render(event):
    wb = openpyxl.Workbook()
    ws = wb.active
    write_sheet(ws, event)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return openpyxl.load_workbook(buf).active


def test_from_swap_sets_highlight():
    ev = Event("劉炆明", "請假", "手動", D(2026, 8, 25), legs=[
        SubLeg("二丁應", "劉炆明", "經濟學", Slot(D(2026, 9, 2), 5), "郭惠茹", from_swap=True),
    ])
    slips = build(ev)
    orig_slip = next(s for s in slips if isinstance(s, TeacherSlip) and s.teacher == "劉炆明")
    assert orig_slip.rows[0].highlight is True

    class_slip = next(s for s in slips if isinstance(s, ClassSlip))
    assert class_slip.rows[0].highlight is True


def test_plain_sub_is_not_highlighted():
    slips = build(samples.get("代課範例"))
    for s in slips:
        for row in s.rows:
            assert getattr(row, "highlight", False) is False


def test_xlsx_has_fill_but_no_j_column():
    # 沒給課表時 J 欄不輸出任何內容（P6 重新啟用 J 欄是條件式的，
    # 只在代課節次對到課表 (兼)/(輔) 標記才會寫，見 test_xlsx_writer.py）
    ws = _render(samples.get("炆明1150831"))
    filled = [c.coordinate for row in ws.iter_rows() for c in row
              if c.fill and c.fill.patternType == "solid"]
    assert filled, "應有反白儲存格"
    assert not [c.value for c in ws["J"] if c.value not in (None, "")]


def test_from_swap_survives_json_round_trip(tmp_path):
    from tiaoke import storage
    from tiaoke.models import Project
    proj = Project(events=[samples.get("炆明1150831")])
    p = tmp_path / "p.json"
    storage.save_project(proj, str(p))
    back = storage.load_project(str(p))
    subs = [l for l in back.events[0].legs if isinstance(l, SubLeg)]
    assert sum(l.from_swap for l in subs) == 3
