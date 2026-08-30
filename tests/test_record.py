import datetime

import openpyxl

from tiaoke import record, samples
from tiaoke.models import Event, Slot, SubLeg, SwapLeg
from tiaoke.ui.controller import AppController

D = datetime.date


def test_semester_and_month_codes():
    assert record.semester_code(D(2026, 9, 2)) == "115-1"
    assert record.semester_code(D(2026, 1, 15)) == "114-1"
    assert record.semester_code(D(2026, 2, 25)) == "114-2"
    assert record.month_code(D(2026, 9, 2)) == "115-09"


def test_sub_leg_makes_one_row():
    ev = Event("吳建勳", "公假", "手動", D(2022, 1, 10), legs=[
        SubLeg("工三", "吳建勳", "食品檢驗", Slot(D(2022, 1, 19), 4), "謝欣瑜"),
    ])
    rows = record.event_to_rows(ev, ts="2026-01-01 00:00")
    assert len(rows) == 1
    r = rows[0]
    assert r[5] == "代課"
    assert r[6] == D(2022, 1, 19)      # 日期
    assert r[8] == 4                    # 節次
    assert r[11] == "吳建勳"            # 原教師
    assert r[13] == "謝欣瑜"            # 實際授課教師
    assert r[15] == "單純代課"          # 型態


def test_swap_leg_makes_two_rows():
    ev = Event("曹朱榜", "公假", "手動", D(2021, 10, 14), legs=[
        SwapLeg("二丁技", "曹朱榜", "彈性學習", Slot(D(2021, 11, 3), 3),
                "張詠竣", "電子學", Slot(D(2021, 11, 5), 5)),
    ])
    rows = record.event_to_rows(ev, ts="x")
    assert [r[5] for r in rows] == ["調課", "調課"]
    # 甲(曹朱榜)換到 11/5 第5節 上 彈性學習
    assert rows[0][6] == D(2021, 11, 5) and rows[0][8] == 5
    assert rows[0][13] == "曹朱榜" and rows[0][10] == "彈性學習"
    # 乙(張詠竣)換到 11/3 第3節 上 電子學
    assert rows[1][6] == D(2021, 11, 3) and rows[1][13] == "張詠竣"


def test_update_record_file_and_stats(tmp_path):
    folder = str(tmp_path)
    ev = samples.get("炆明1150831")
    rep = record.update_record(folder, ev)
    assert rep.ok
    assert rep.path.endswith("115-1調代課記錄.xlsx")
    assert rep.added_sub == 4 and rep.added_swap == 6

    wb = openpyxl.load_workbook(rep.path)
    assert wb.sheetnames[:2] == ["調代課明細", "月統計"]
    detail = list(wb["調代課明細"].iter_rows(min_row=2, values_only=True))
    assert len(detail) == 10

    # 月統計：哪幾列該出現（堂數本身是 Excel 公式）
    keys = record.stat_keys(wb["調代課明細"])
    assert ("115-09", "郭惠茹") in keys
    assert ("115-09", "劉炆明") in keys
    assert ("115-08", "劉炆明") in keys           # 8/31 那堂歸 8 月
    assert keys[("115-09", "郭惠茹")]["dates"].count(datetime.date(2026, 9, 2)) == 2

    stats = {(r[0].value, r[1].value): r for r in wb["月統計"].iter_rows(min_row=2)}
    # 堂數欄是 COUNTIFS 公式
    f = stats[("115-09", "郭惠茹")][3].value
    assert f.startswith("=COUNTIFS(") and "代課" in f and "$B" in f
    assert "DATE(2026,9,1)" in f and "DATE(2026,10,1)" in f


def test_stat_keys_dates_are_actual_class_dates():
    """月統計的『代課日期明細』依實際上課日歸月（不是公告日）。"""
    from openpyxl import Workbook
    wb = Workbook()
    ev = samples.get("炆明1150831")   # 公告日 8/25，但代課發生在 8/31 與 9/2
    record._write_detail(wb, record.event_to_rows(ev, ts="x"))
    keys = record.stat_keys(wb["調代課明細"])

    # 郭惠茹 9/2 代課 2 堂 → 9 月那列的 dates 有兩個 9/2
    assert keys[("115-09", "郭惠茹")]["dates"] == [datetime.date(2026, 9, 2)] * 2
    # 徐惠珠 8/31 代課 1 堂
    assert keys[("115-08", "徐惠珠")]["dates"] == [datetime.date(2026, 8, 31)]
    # 沒有任何一列的月份是公告月以外的錯月（都在 8 或 9 月）
    assert {m.split("-")[1] for (m, _t) in keys} <= {"08", "09"}


def test_regenerate_replaces_not_duplicates(tmp_path):
    folder = str(tmp_path)
    ev = samples.get("代課範例")
    r1 = record.update_record(folder, ev)
    r2 = record.update_record(folder, ev)
    assert r2.removed == r1.added_sub + r1.added_swap
    wb = openpyxl.load_workbook(r2.path)
    detail = [x for x in wb["調代課明細"].iter_rows(min_row=2, values_only=True) if x[0]]
    assert len(detail) == 1


def test_controller_writes_record_on_generate(tmp_path):
    c = AppController()
    c.record_folder = str(tmp_path / "rec")
    c.new_event()
    c.update_event_fields(originator="余瑞文", form_no="手動", announce_date=D(2026, 8, 25))
    c.add_sub_leg(klass="電機一", orig_teacher="余瑞文", subject="健康與護理",
                  date=D(2026, 9, 1), period=2, sub_teacher="王小明")
    c.generate(to_master=False, save_new=True, dest_path=str(tmp_path / "x.xlsx"))
    assert c.last_record is not None and c.last_record.ok
    assert c.last_record.added_sub == 1
