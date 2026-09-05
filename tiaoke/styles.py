"""通知單樣式常數與 openpyxl helper。數值取自範例檔。"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

FONT_NAME = "標楷體"
BODY_SIZE = 12
TITLE_SIZE = 18

# 欄寬取自範例檔「範例1」分頁（原始精確值）
COL_WIDTHS = {
    "A": 12.26953125, "B": 9.0, "C": 6.90625, "D": 7.0, "E": 13.0,
    "F": 8.6328125, "G": 6.08984375, "H": 6.6328125, "I": 17.90625,
}
# 列高取自「範例1」：橫幅列 25、教師單資料列 34、班級單資料列 ~22
TITLE_ROW_H = 25.0
TEACHER_ROW_H = 34.0
CLASS_ROW_H = 22.0
DATA_ROW_H = TEACHER_ROW_H  # 舊名相容
NOTE_ROW_H = 72.0

DATE_FMT = 'm"月"d"日"'

# 班級標題式（title）用的置中大標
CLASS_TITLE_TEXT = "             班　級　調　代　課　通　知　單"

PRINT_SCALE = 95
A4_PAPER = 9

_THIN = Side(style="thin")
_NONE = Side(style=None)

CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
NOTE_ALIGN = Alignment(vertical="center", wrap_text=True)

# 「先調後代」列的反白（淡金色，列印仍清楚）
HIGHLIGHT_FILL = PatternFill("solid", fgColor="FFF2CC")


def body_font(bold: bool = False, color: str | None = None) -> Font:
    return Font(name=FONT_NAME, size=BODY_SIZE, bold=bold, color=color)


def title_font() -> Font:
    return Font(name=FONT_NAME, size=TITLE_SIZE, bold=True)


def mark_font() -> Font:
    """J 欄「兼課」／「八」註記：藍字、標楷、粗體。"""
    return Font(name=FONT_NAME, size=BODY_SIZE, bold=True, color="0000FF")


def box(l=False, r=False, t=False, b=False) -> Border:
    return Border(
        left=_THIN if l else _NONE,
        right=_THIN if r else _NONE,
        top=_THIN if t else _NONE,
        bottom=_THIN if b else _NONE,
    )


ALL_BOX = box(True, True, True, True)
BOTTOM_ONLY = box(b=True)


def put(ws: Worksheet, coord: str, value, *, font: Font | None = None,
        align: Alignment | None = None, border: Border | None = None,
        fmt: str | None = None, fill: PatternFill | None = None):
    """寫入單一儲存格並套用樣式。"""
    cell = ws[coord]
    cell.value = value
    cell.font = font or body_font()
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border
    if fmt is not None:
        cell.number_format = fmt
    if fill is not None:
        cell.fill = fill
    return cell


def outline_grid(ws: Worksheet, cell_range: str) -> None:
    """把整個矩形範圍的每一格都畫上四邊細框線（在合併之前呼叫）。"""
    min_c, min_r, max_c, max_r = range_boundaries(cell_range)
    for row in range(min_r, max_r + 1):
        for col in range(min_c, max_c + 1):
            ws.cell(row=row, column=col).border = ALL_BOX


def set_col_widths(ws: Worksheet) -> None:
    for col, width in COL_WIDTHS.items():
        # 指定 width 後 openpyxl 的 customWidth 會自動回報 True
        ws.column_dimensions[col].width = width
