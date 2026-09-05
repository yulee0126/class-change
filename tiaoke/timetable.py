"""讀入「教師課表」PDF，解析成每位老師的每週課表，存成 JSON 供表單下拉使用。

PDF 版型（國立關西高中 115-1「暫行」課表，一頁一位老師）：
  · 標題列：學校、學年期
  · 「教 師：姓名 (編號) 職稱 頁數：NNN」
  · 週課表格：節次 1–8 × 星期一~五；每格 = 科目(可跨行) / 授課班級 / 上課地點(可省)
  · 「本表自115年 MM 月 DD 日起至115年 MM 月 DD日止」

領域（學習領域）在此 PDF 未提供，欄位保留為空字串。
"""

from __future__ import annotations

import datetime
import re
from collections import Counter
from dataclasses import asdict, dataclass, field

_FULL2HALF = str.maketrans("０１２３４５６７８９：", "0123456789:")
_CIRCLED = "０１２３４５６７８９"
_WEEKDAYS = ["星期一", "星期二", "星期三", "星期四", "星期五"]

# 沒有授課班級、且屬於下列名稱者＝會議／非課務，略過不收
_SKIP_IF_NO_CLASS = {
    "行政會報", "導師會報", "教學研究會", "校務會議", "朝會", "升旗", "降旗",
    "午休", "打掃", "掃地", "課發會", "領域會議", "處務會議",
}

FORMAT = 1


# --------------------------------------------------------------------------
# 資料結構
# --------------------------------------------------------------------------

@dataclass
class Slot:
    weekday: int            # 1=星期一 … 5=星期五
    period: int             # 1–8
    subject: str
    klass: str
    location: str = ""
    domain: str = ""        # 學習領域（此 PDF 無，保留空）
    note: str = ""          # 例：(彈性全學期)、(兼)＝兼課、(輔)＝輔導
    co_teachers: list[str] = field(default_factory=list)  # 這節課協同教學的其他老師（教師配當表比對出來的）


@dataclass
class TeacherTable:
    name: str
    tid: str = ""           # 教師編號，例 "00101"
    title: str = ""         # 職稱，例 "園藝科主任"
    slots: list[Slot] = field(default_factory=list)

    def on(self, weekday: int) -> list[Slot]:
        return sorted((s for s in self.slots if s.weekday == weekday),
                      key=lambda s: s.period)

    # ---- 課表校對 GUI 用：一格＝一個星期＋節次（合班課＝多個 Slot 共用同一格）----
    def slot_group(self, weekday: int, period: int) -> list[Slot]:
        return [s for s in self.slots if s.weekday == weekday and s.period == period]

    def set_slot_group(self, weekday: int, period: int, subject: str,
                       klasses: list[str], location: str = "", note: str = "",
                       co_teachers: list[str] | None = None) -> None:
        """整格改寫：先刪同格舊資料，再依 klasses 逐一新增（多班＝合班課）。"""
        self.slots = [s for s in self.slots
                      if not (s.weekday == weekday and s.period == period)]
        for k in (klasses or [""]):
            self.slots.append(Slot(weekday=weekday, period=period, subject=subject,
                                   klass=k, location=location, note=note,
                                   co_teachers=list(co_teachers or [])))
        self.slots.sort(key=lambda s: (s.weekday, s.period, s.klass))

    def delete_slot_group(self, weekday: int, period: int) -> None:
        self.slots = [s for s in self.slots
                      if not (s.weekday == weekday and s.period == period)]


@dataclass
class Timetable:
    school: str = ""
    semester: str = ""
    valid_from: str = ""    # ISO
    valid_to: str = ""
    source_pdf: str = ""
    period_times: dict[int, list[str]] = field(default_factory=dict)  # period -> [start, end]
    teachers: dict[str, TeacherTable] = field(default_factory=dict)

    # ---- 查詢 ----
    def teacher_names(self) -> list[str]:
        return sorted(self.teachers)

    def slots_for(self, teacher: str, weekday: int) -> list[Slot]:
        t = self.teachers.get(teacher.strip())
        return t.on(weekday) if t else []

    # ---- 序列化 ----
    def to_dict(self) -> dict:
        return {
            "format": FORMAT,
            "school": self.school,
            "semester": self.semester,
            "valid_from": self.valid_from,
            "valid_to": self.valid_to,
            "source_pdf": self.source_pdf,
            "period_times": {str(k): v for k, v in self.period_times.items()},
            "teachers": {
                name: {
                    "name": t.name, "tid": t.tid, "title": t.title,
                    "slots": [asdict(s) for s in t.slots],
                }
                for name, t in self.teachers.items()
            },
        }

    @classmethod
    def from_dict(cls, d: dict) -> "Timetable":
        tt = cls(
            school=d.get("school", ""), semester=d.get("semester", ""),
            valid_from=d.get("valid_from", ""), valid_to=d.get("valid_to", ""),
            source_pdf=d.get("source_pdf", ""),
            period_times={int(k): list(v) for k, v in d.get("period_times", {}).items()},
        )
        for name, td in d.get("teachers", {}).items():
            tt.teachers[name] = TeacherTable(
                name=td.get("name", name), tid=td.get("tid", ""), title=td.get("title", ""),
                slots=[Slot(**s) for s in td.get("slots", [])],
            )
        return tt


# --------------------------------------------------------------------------
# 解析
# --------------------------------------------------------------------------

def parse_pdf(path: str) -> Timetable:
    import pdfplumber

    tt = Timetable(source_pdf=path)
    pages_cells: list[tuple[str, str, str, dict, list]] = []

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            meta = _parse_meta(text)
            if not tt.school:
                tt.school = meta.get("school", "")
                tt.semester = meta.get("semester", "")
            if not tt.valid_from and meta.get("valid_from"):
                tt.valid_from = meta["valid_from"]
                tt.valid_to = meta["valid_to"]

            tables = page.extract_tables()
            grid = tables[0] if tables else []
            _collect_period_times(grid, tt.period_times)
            cells = _grid_cells(grid)  # list[(period, weekday, [lines])]
            if meta.get("name"):
                pages_cells.append((meta["name"], meta.get("tid", ""),
                                    meta.get("title", ""), None, cells))

    class_vocab, room_vocab = _build_vocab(pages_cells)

    for name, tid, title, _n, cells in pages_cells:
        t = tt.teachers.setdefault(name, TeacherTable(name=name, tid=tid, title=title))
        if tid and not t.tid:
            t.tid = tid
        if title and not t.title:
            t.title = title
        seen: set[tuple] = set()
        for period, weekday, lines in cells:
            parsed = _parse_cell(lines, class_vocab, room_vocab)
            if not parsed:
                continue
            classes = parsed["classes"] or [""]
            note = parsed["note"]
            # 校內慣例：第八節一律算輔導，不論 PDF 該格是否印出 (輔) 字樣。
            if period == 8 and "(輔)" not in note:
                note += "(輔)"
            for kls in classes:
                key = (weekday, period, parsed["subject"], kls)
                if key in seen:
                    continue
                seen.add(key)
                t.slots.append(Slot(
                    weekday=weekday, period=period,
                    subject=parsed["subject"], klass=kls,
                    location=parsed["location"], note=note,
                ))
        t.slots.sort(key=lambda s: (s.weekday, s.period, s.klass))

    return tt


# ---- metadata ----

def _parse_meta(text: str) -> dict:
    m: dict = {}
    for line in text.splitlines():
        s = line.strip()
        if "學年度" in s and "教師課表" in s:
            m["school"] = s.split()[0]
            mm = re.search(r"(\d+學年度第[一二三])", s.replace(" ", ""))
            if mm:
                m["semester"] = mm.group(1)
        if s.startswith("教") and "師：" in s:
            mm = re.search(r"師：\s*([^\s(（]+)\s*[（(]([0-9A-Za-z]+)[)）]\s*([^\s]*)", s)
            if mm:
                m["name"] = mm.group(1).strip()
                m["tid"] = mm.group(2).strip()
                title = mm.group(3).strip()
                if title and not title.startswith("頁數"):
                    m["title"] = title
        mm = re.search(r"本表自(\d+)年\s*(\d+)\s*月\s*(\d+)\s*日起至(\d+)年\s*(\d+)\s*月\s*(\d+)\s*日", s)
        if mm:
            y1, m1, d1, y2, m2, d2 = (int(x) for x in mm.groups())
            m["valid_from"] = datetime.date(y1 + 1911, m1, d1).isoformat()
            m["valid_to"] = datetime.date(y2 + 1911, m2, d2).isoformat()
    return m


def _collect_period_times(grid: list, out: dict) -> None:
    for row in grid:
        if not row or len(row) < 3:
            continue
        pnum = _period_num(row[1])
        if pnum is None or pnum in out:
            continue
        raw = (row[2] or "").translate(_FULL2HALF)
        times = re.findall(r"\d{1,2}:\d{2}", raw)
        if len(times) >= 2:
            out[pnum] = [times[0], times[1]]


def _period_num(cell) -> int | None:
    if not cell:
        return None
    s = str(cell).strip().translate(_FULL2HALF)
    if s.isdigit() and 1 <= int(s) <= 9:
        return int(s)
    return None


def _grid_cells(grid: list) -> list:
    """回傳 [(period, weekday, [lines]), ...]。"""
    cells = []
    for row in grid:
        if not row or len(row) < 8:
            continue
        pnum = _period_num(row[1])
        if pnum is None:
            continue
        for di, raw in enumerate(row[3:8], start=1):
            if not raw:
                continue
            lines = [x.strip() for x in str(raw).split("\n") if x.strip()]
            if lines:
                cells.append((pnum, di, lines))
    return cells


# ---- 詞彙表（兩遍解析）----

def _build_vocab(pages_cells: list) -> tuple[set, set]:
    last_tokens: Counter = Counter()
    for *_x, cells in pages_cells:
        for _p, _w, lines in cells:
            if len(lines) >= 2:
                for tok in (lines[-1], lines[-2]):
                    tok, _mark = _strip_mark(tok)
                    if tok:
                        last_tokens[tok] += 1

    classes, rooms = set(), set()
    for tok, cnt in last_tokens.items():
        if _room_like(tok):
            rooms.add(tok)
        elif _class_like(tok):
            classes.add(tok)
    return classes, rooms


_CLASS_RE = re.compile(
    r"^(高[一二三]|國[一二三])?[一-鿿]{0,3}[一二三四五六甲乙丙丁戊][技應餐商甲乙丙丁]?$")

# 課名／班級後常黏著的鐘點標記：(兼)=兼課、(輔)=輔導（對應頁尾「基本鐘點／兼課／輔導」統計）。
# PDF 版面換行時，標記可能單獨成一行，也可能直接黏在科目或班級文字後面。
_MARK_RE = re.compile(r"[（(]\s*(兼|輔)\s*[）)]\s*$")


def _strip_mark(line: str) -> tuple[str, str | None]:
    """去掉行尾黏著的 (兼)/(輔) 標記，回傳 (去除後文字, 標記或 None)。"""
    m = _MARK_RE.search(line)
    if not m:
        return line, None
    return line[:m.start()].rstrip(), m.group(1)


def _class_like(tok: str) -> bool:
    if not (2 <= len(tok) <= 6):
        return False
    if _room_like(tok):
        return False
    if tok[-1] not in "一二三四五六甲乙丙丁戊技應餐商":
        return False
    if tok[-1] in "技應餐商" and len(tok) < 3:
        return False
    # 排除明顯是科目結尾的
    if tok.endswith(("實習", "概論", "實務", "作", "學", "論", "習作")):
        return False
    return bool(_CLASS_RE.match(tok))


def _room_like(tok: str) -> bool:
    return any(k in tok for k in (
        "教室", "工廠", "實驗", "教學", "專業實習", "空間", "洗車", "球場",
        "電腦", "視聽", "自主學", "加工一廠", "加工二廠", "畜牧場", "造園場",
        "配線場", "實習場", "工場", "會議室", "辦公室",
    )) or bool(re.search(r"[0-9A-Za-z]F", tok)) or tok.endswith(
        ("樓", "廠", "館", "場", "室"))


def _is_class(tok: str, class_vocab: set) -> bool:
    return tok in class_vocab or _class_like(tok)


def _is_room(tok: str, room_vocab: set) -> bool:
    return tok in room_vocab or _room_like(tok)


def _split_slash(tok: str) -> list[str]:
    return [p.strip() for p in re.split(r"[/／、]", tok) if p.strip()]


def _all_classes(tok: str, vocab: set) -> bool:
    parts = _split_slash(tok)
    return len(parts) >= 2 and all(_is_class(p, vocab) for p in parts)


def _parse_cell(lines: list[str], class_vocab: set, room_vocab: set) -> dict | None:
    raw = [l.strip() for l in lines if l.strip()]
    if not raw:
        return None

    marks: list[str] = []
    stripped: list[str] = []
    for l in raw:
        cleaned, mark = _strip_mark(l)
        if mark:
            marks.append(f"({mark})")
        if cleaned:
            stripped.append(cleaned)
    raw = stripped
    if not raw:
        return None

    notes = [l for l in raw if l.startswith(("(", "（"))]
    raw = [l for l in raw if l not in notes]
    if not raw:
        return None
    notes = marks + notes

    # 合班列常跨行斷開（'加工一/商' + '經一/餐飲一/'）→ 先接回來
    merged: list[str] = []
    i = 0
    while i < len(raw):
        cur = raw[i]
        while ("/" in cur or "／" in cur) and not _all_classes(cur, class_vocab) \
                and i + 1 < len(raw):
            i += 1
            cur += raw[i]
        merged.append(cur)
        i += 1

    # 把合班列展開成多個 token
    flat: list[str] = []
    for l in merged:
        if _all_classes(l, class_vocab):
            flat.extend(_split_slash(l))
        else:
            flat.append(l)
    lines = flat

    cls_pos = [i for i, l in enumerate(lines) if _is_class(l, class_vocab)]
    if cls_pos:
        start, end = cls_pos[0], cls_pos[-1]
        subject = "".join(lines[:start])
        classes = [lines[i] for i in range(start, end + 1) if _is_class(lines[i], class_vocab)]
        location = "".join(lines[end + 1:])
    else:
        loc_parts: list[str] = []
        while len(lines) > 1 and _is_room(lines[-1], room_vocab):
            loc_parts.insert(0, lines.pop())
        subject = "".join(lines)
        classes = []
        location = "".join(loc_parts)

    subject = subject.strip()
    if not subject:
        return None
    if not classes and subject in _SKIP_IF_NO_CLASS:
        return None
    seen: set[str] = set()
    classes = [c for c in classes if not (c in seen or seen.add(c))]
    return {
        "subject": subject,
        "classes": classes,
        "location": location.strip(),
        "note": "".join(notes),
    }


# --------------------------------------------------------------------------
# 協同教學（教師配當表）
# --------------------------------------------------------------------------
#
# 「教師配當表」逐老師逐課程列出授課班級／課程名稱／學分數，其中「協同」欄標記
# 兩位（以上）老師一起教同一堂課。這份表沒有星期／節次，只能靠跟已匯入的
# 教師課表（PDF）比對「同班級、同課程、同星期節次」來確認真的是協同授課
# （而不是同名課程但根本是不同時段各自上課，例如同一堂「專題實作」很多老師
# 各自帶不同組別）。班級代碼兩邊常不一樣（配當表用簡稱如「職二」，課表用
# 全稱如「綜職二」），用「簡稱各字依序出現在全稱裡」判斷是否同一班。

_CO_TEACH_MARK = "協同"  # 涵蓋「協同」「專案協同」


def parse_co_teaching_xlsx(path: str) -> list[tuple[str, str, str, str]]:
    """讀教師配當表，回傳 [(教師, 授課班級簡稱, 課程名稱, 標記), ...]。

    標記可能是 "協同"／"專案協同"／"分組"／其他代碼／空字串；
    是否視為協同教學交給 `apply_co_teaching` 判斷。
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = _find_peidang_sheet(wb)

    rows: list[tuple[str, str, str, str]] = []
    cur_teacher = ""
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 7:
            continue
        name, klass, subject, mark = row[2], row[3], row[4], row[6]
        if name:
            cur_teacher = str(name).strip()
        if not (cur_teacher and klass and subject):
            continue
        rows.append((cur_teacher, str(klass).strip(), str(subject).strip(),
                    str(mark).strip() if mark else ""))
    return rows


def _find_peidang_sheet(wb):
    """找表頭含「教師姓名」「授課班級」「課程名稱」的分頁。

    有些配當表會把舊學期的表留在另一個分頁、表頭長得一模一樣（只是「課程科別」
    欄位在新表已經沒在用，協同/分組標記改擠到隔壁欄），所以不能只憑表頭判斷，
    要選「協同」「分組」標記出現次數最多的那個分頁，才是目前學期真正在用的。
    """
    candidates = []
    for ws in wb.worksheets:
        header_ok = any(
            row and "教師姓名" in row and "授課班級" in row and "課程名稱" in row
            for row in ws.iter_rows(min_row=1, max_row=3, values_only=True)
        )
        if not header_ok:
            continue
        mark_count = sum(
            1 for row in ws.iter_rows(min_row=3, values_only=True)
            for v in row if isinstance(v, str) and ("協同" in v or "分組" in v)
        )
        candidates.append((mark_count, ws))
    if not candidates:
        return wb.worksheets[-1]
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


def _klass_matches(short: str, full: str) -> bool:
    """配當表簡稱（'職二'）跟課表全稱（'綜職二'）是否同一班：
    簡稱各字依序（可不連續）出現在全稱裡即可。"""
    it = iter(full)
    return all(ch in it for ch in short)


def apply_co_teaching(timetable: Timetable, rows: list[tuple[str, str, str, str]]) -> int:
    """用配當表資料比對 timetable 裡的實際節次，把確認協同的 Slot 標上 co_teachers。

    規則：同一 (班級簡稱, 課程名稱) 至少 2 位不同老師、且至少一邊標了「(專案)協同」，
    才去查課表；查到同班級全稱、同星期、同節次有 2 位以上老師才算數（見模組說明）。
    回傳被標記／更新的 Slot 數。
    """
    groups: dict[tuple[str, str], list[tuple[str, str]]] = {}
    for teacher, klass_short, subject, mark in rows:
        groups.setdefault((klass_short, subject), []).append((teacher, mark))

    touched = 0
    for (klass_short, subject), entries in groups.items():
        distinct_teachers = {t for t, _m in entries}
        if len(distinct_teachers) < 2:
            continue
        if not any(_CO_TEACH_MARK in mark for _t, mark in entries):
            continue

        by_slot: dict[tuple, list[tuple[str, Slot]]] = {}
        for teacher in distinct_teachers:
            t = timetable.teachers.get(teacher)
            if not t:
                continue
            for s in t.slots:
                if s.subject == subject and _klass_matches(klass_short, s.klass):
                    by_slot.setdefault((s.klass, s.weekday, s.period), []).append((teacher, s))

        for occs in by_slot.values():
            names = sorted({teacher for teacher, _s in occs})
            if len(names) < 2:
                continue
            for teacher, s in occs:
                others = [n for n in names if n != teacher]
                if s.co_teachers != others:
                    s.co_teachers = others
                    touched += 1
    return touched
