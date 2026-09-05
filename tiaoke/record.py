"""調代課記錄檔：每學期一個 xlsx，產生通知單時同步寫入，供每月鐘點費統計。

工作表：
  · 調代課明細 —— 一列＝一個異動節次（代課 1 列、調課 2 列）；依建立順序往下排
  · 月統計   —— 一列＝一位老師一個月（長表），程式每次自動重算
"""

from __future__ import annotations

import datetime
import os
from collections import Counter, OrderedDict
from dataclasses import dataclass, field

from openpyxl import Workbook, load_workbook

from . import roc
from .builder import lookup_tags
from .models import CoSwapLeg, Event, SubLeg, SwapLeg

DETAIL_SHEET = "調代課明細"
STATS_SHEET = "月統計"

DETAIL_HEADERS = [
    "記錄時間", "事件", "假單編號", "假別", "發起教師", "類型", "日期", "星期",
    "節次", "班級", "授課科目", "原教師", "原教師編號", "實際授課教師", "實際教師編號",
    "型態", "公告日期", "代課別", "鐘點費", "備註",
]
STATS_HEADERS = ["月份", "教師", "教師編號", "代課堂數", "代課(兼課)堂數", "代課(第八節)堂數",
                 "被代堂數", "調課堂數", "代課日期（民國）"]

_DATE_FMT = "yyyy/mm/dd"


# --------------------------------------------------------------------------

@dataclass
class RecordReport:
    path: str = ""
    added_sub: int = 0        # 新增的代課列數
    added_swap: int = 0       # 新增的調課列數
    removed: int = 0          # 刪掉的舊列數（同一事件）
    error: str = ""

    @property
    def ok(self) -> bool:
        return not self.error

    def __str__(self) -> str:
        if self.error:
            return f"記錄檔失敗：{self.error}"
        return f"代課 {self.added_sub} 筆、調課 {self.added_swap} 筆" + (
            f"（覆蓋舊 {self.removed} 筆）" if self.removed else "")


# --------------------------------------------------------------------------
# 學期 / 月份
# --------------------------------------------------------------------------

def semester_code(d: datetime.date) -> str:
    """民國學年-學期：8月~隔年1月＝上(1)、2~7月＝下(2)。例 2026-09 → '115-1'。"""
    if d.month >= 8:
        return f"{d.year - 1911}-1"
    if d.month == 1:
        return f"{d.year - 1911 - 1}-1"
    return f"{d.year - 1911 - 1}-2"


def month_code(d: datetime.date) -> str:
    return f"{d.year - 1911}-{d.month:02d}"


def _roc_ymd(d: datetime.date) -> str:
    return f"{d.year - 1911}-{d.month:02d}-{d.day:02d}"


def record_path(folder: str, ref_date: datetime.date) -> str:
    return os.path.join(folder, f"{semester_code(ref_date)}調代課記錄.xlsx")


# --------------------------------------------------------------------------
# Event → 明細列
# --------------------------------------------------------------------------

def _tid(timetable, name: str) -> str:
    if not timetable:
        return ""
    t = timetable.teachers.get((name or "").strip())
    return t.tid if t else ""


def event_to_rows(event: Event, timetable=None, ts: str | None = None) -> list[list]:
    """把一個事件展開成明細列（每列是對應 DETAIL_HEADERS 的 list）。"""
    ts = ts or datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    rows: list[list] = []

    def row(kind, d, period, klass, subject, orig, actual, mode, note="", sub_type=""):
        return [
            ts, event.sheet_name, event.form_no, event.leave_type, event.originator,
            kind, d, roc.weekday_cn(d), int(period), klass, subject,
            orig, _tid(timetable, orig), actual, _tid(timetable, actual),
            mode, event.announce_date, sub_type, "", note,
        ]

    def _sub_type(leg: SubLeg) -> str:
        is_extra, is_period8 = lookup_tags(timetable, leg.orig_teacher, leg.slot)
        tags = []
        if is_extra:
            tags.append("兼課")
        if is_period8:
            tags.append("第八節")
        return "、".join(tags)

    for leg in event.legs:
        if isinstance(leg, SubLeg):
            rows.append(row(
                "代課", leg.slot.date, leg.slot.period, leg.klass, leg.subject,
                leg.orig_teacher, leg.sub_teacher,
                "先調後代" if leg.from_swap else "單純代課",
                sub_type=_sub_type(leg),
            ))
        elif isinstance(leg, SwapLeg):
            # 甲換到乙的時段（slot_b）上甲的科目
            rows.append(row(
                "調課", leg.slot_b.date, leg.slot_b.period, leg.klass, leg.subject_a,
                leg.teacher_b, leg.teacher_a, "",
                note=f"原 {roc.slot_label(leg.slot_a.date, leg.slot_a.period)}",
            ))
            # 乙換到甲的時段（slot_a）上乙的科目
            rows.append(row(
                "調課", leg.slot_a.date, leg.slot_a.period, leg.klass, leg.subject_b,
                leg.teacher_a, leg.teacher_b, "",
                note=f"原 {roc.slot_label(leg.slot_b.date, leg.slot_b.period)}",
            ))
        elif isinstance(leg, CoSwapLeg):
            # A 側每位老師各一列：換到 B 側時段（slot_b）上 A 側的科目
            a_label = "、".join(leg.teachers_a)
            b_label = "、".join(leg.teachers_b)
            for t in leg.teachers_a:
                rows.append(row(
                    "調課", leg.slot_b.date, leg.slot_b.period, leg.klass, leg.subject_a,
                    b_label, t, "",
                    note=f"原 {roc.slot_label(leg.slot_a.date, leg.slot_a.period)}",
                ))
            # B 側每位老師各一列：換到 A 側時段（slot_a）上 B 側的科目
            for t in leg.teachers_b:
                rows.append(row(
                    "調課", leg.slot_a.date, leg.slot_a.period, leg.klass, leg.subject_b,
                    a_label, t, "",
                    note=f"原 {roc.slot_label(leg.slot_b.date, leg.slot_b.period)}",
                ))
    return rows


# --------------------------------------------------------------------------
# 寫檔
# --------------------------------------------------------------------------

def update_record(folder: str, event: Event, timetable=None) -> RecordReport:
    try:
        os.makedirs(folder, exist_ok=True)
        ref_date = event.effective_sheet_date
        path = record_path(folder, ref_date)
        new_rows = event_to_rows(event, timetable)

        wb = load_workbook(path) if os.path.exists(path) else Workbook()
        if wb.sheetnames == ["Sheet"]:
            wb.remove(wb.active)

        existing = _read_detail(wb)
        kept = [r for r in existing if (r[1] or "") != event.sheet_name]
        removed = len(existing) - len(kept)

        _write_detail(wb, kept + new_rows)
        _rebuild_stats(wb)
        _force_recalc(wb)
        wb.save(path)

        return RecordReport(
            path=path,
            added_sub=sum(1 for r in new_rows if r[5] == "代課"),
            added_swap=sum(1 for r in new_rows if r[5] == "調課"),
            removed=removed,
        )
    except PermissionError:
        return RecordReport(error="記錄檔可能正在 Excel 中開啟，請關閉後再產生一次。")
    except Exception as exc:  # noqa: BLE001
        return RecordReport(error=str(exc))


def _read_detail(wb: Workbook) -> list[list]:
    if DETAIL_SHEET not in wb.sheetnames:
        return []
    ws = wb[DETAIL_SHEET]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and any(c is not None and c != "" for c in r):
            row = list(r) + [None] * (len(DETAIL_HEADERS) - len(r))
            out.append(row[:len(DETAIL_HEADERS)])
    return out


def _write_detail(wb: Workbook, rows: list[list]) -> None:
    if DETAIL_SHEET in wb.sheetnames:
        del wb[DETAIL_SHEET]
    ws = wb.create_sheet(DETAIL_SHEET, 0)
    ws.append(DETAIL_HEADERS)
    for row in rows:
        ws.append(row)
    for cell in ws["G"][1:]:      # 日期
        cell.number_format = _DATE_FMT
    for cell in ws["Q"][1:]:      # 公告日期
        cell.number_format = _DATE_FMT
    widths = {"A": 16, "B": 14, "F": 6, "G": 12, "K": 14, "L": 10, "N": 12, "Q": 12, "T": 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def _as_date(v) -> datetime.date | None:
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, str):
        try:
            return datetime.date.fromisoformat(v.strip()[:10])
        except ValueError:
            return None
    return None


def stat_keys(detail_ws) -> "OrderedDict[tuple, dict]":
    """掃「調代課明細」，回傳需要在月統計出現的 (月份, 教師) → {tid, 代課日期清單}。
    堂數本身交給 Excel 公式算（改明細會即時更新），這裡只決定要列哪幾列。"""
    out: "OrderedDict[tuple, dict]" = OrderedDict()

    def touch(month: str, teacher: str, tid: str) -> dict:
        key = (month, (teacher or "").strip())
        b = out.setdefault(key, {"tid": "", "dates": []})
        b["tid"] = b["tid"] or (tid or "")
        return b

    for r in detail_ws.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        kind, d = r[5], _as_date(r[6])
        if d is None:
            continue
        month = month_code(d)
        orig, orig_tid, actual, actual_tid = r[11], r[12], r[13], r[14]
        if kind == "代課":
            touch(month, actual, actual_tid)["dates"].append(d)
            touch(month, orig, orig_tid)
        elif kind == "調課":
            touch(month, actual, actual_tid)

    return OrderedDict((k, v) for k, v in out.items() if k[1])


def _mc_ym(month_code_str: str) -> tuple[int, int]:
    acad, m = month_code_str.split("-")
    return int(acad) + 1911, int(m)


def _rebuild_stats(wb: Workbook) -> None:
    keys = stat_keys(wb[DETAIL_SHEET])

    if STATS_SHEET in wb.sheetnames:
        del wb[STATS_SHEET]
    ws = wb.create_sheet(STATS_SHEET)
    ws.append(STATS_HEADERS)
    q = f"'{DETAIL_SHEET}'"

    for i, ((month, teacher), v) in enumerate(sorted(keys.items()), start=2):
        y, m = _mc_ym(month)
        ny, nm = (y + 1, 1) if m == 12 else (y, m + 1)
        span = (f"{q}!$G:$G,\">=\"&DATE({y},{m},1),"
                f"{q}!$G:$G,\"<\"&DATE({ny},{nm},1)")
        cnt = Counter(_roc_ymd(d) for d in v["dates"])
        dates_str = "、".join(f"{k}×{n}" if n > 1 else k for k, n in sorted(cnt.items()))
        ws.append([
            month, teacher, v["tid"],
            f'=COUNTIFS({q}!$F:$F,"代課",{q}!$N:$N,$B{i},{span})',
            f'=COUNTIFS({q}!$F:$F,"代課",{q}!$N:$N,$B{i},{q}!$R:$R,"*兼課*",{span})',
            f'=COUNTIFS({q}!$F:$F,"代課",{q}!$N:$N,$B{i},{q}!$R:$R,"*第八節*",{span})',
            f'=COUNTIFS({q}!$F:$F,"代課",{q}!$L:$L,$B{i},{span})',
            f'=COUNTIFS({q}!$F:$F,"調課",{q}!$N:$N,$B{i},{span})',
            dates_str,
        ])

    ws["K1"] = "堂數為公式，改「調代課明細」會自動更新；「代課(兼課)/(第八節)堂數」統計的是\n代課老師代的那節，原老師課表若標了 (兼)/(輔)；「代課日期（民國）」是產生當下的快照。"
    ws["K1"].font = _hint_font()
    ws.freeze_panes = "A2"
    for col, w in {"A": 9, "B": 10, "C": 10, "E": 14, "F": 16, "I": 22}.items():
        ws.column_dimensions[col].width = w


def _hint_font():
    from openpyxl.styles import Font
    return Font(size=9, italic=True, color="808080")


def _force_recalc(wb: Workbook) -> None:
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass


def build_report(rows: list[list]) -> Workbook:
    """把已收集好的明細列組成一份全新的記錄檔活頁簿（供「產製報表」用，不覆蓋既有檔案）。"""
    wb = Workbook()
    wb.remove(wb.active)
    _write_detail(wb, rows)
    _rebuild_stats(wb)
    _force_recalc(wb)
    return wb


def rebuild_stats_file(path: str) -> RecordReport:
    """只重算某個記錄檔的「月統計」（使用者手改明細後用）。"""
    try:
        if not os.path.exists(path):
            return RecordReport(error=f"找不到記錄檔：{path}")
        wb = load_workbook(path)
        if DETAIL_SHEET not in wb.sheetnames:
            return RecordReport(error="這個檔沒有「調代課明細」。")
        _rebuild_stats(wb)
        _force_recalc(wb)
        wb.save(path)
        return RecordReport(path=path)
    except PermissionError:
        return RecordReport(error="記錄檔可能正在 Excel 中開啟。")
    except Exception as exc:  # noqa: BLE001
        return RecordReport(error=str(exc))
