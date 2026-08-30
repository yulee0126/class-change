"""輸出動作：寫入總表 / 另存新檔。兩者可於同一次呼叫一起執行。"""

from __future__ import annotations

import os
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook

from . import xlsx_writer
from .models import Event


@dataclass
class TargetResult:
    target: str          # "master" | "new"
    path: str
    ok: bool
    replaced_sheet: bool = False
    last_row: int = 0
    error: str = ""


def build_single_workbook(event: Event) -> Workbook:
    """只含該事件一張工作表的新活頁簿。"""
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_title(event.sheet_name)
    xlsx_writer.write_sheet(ws, event)
    return wb


def save_as_new(event: Event, dest_path: str) -> TargetResult:
    try:
        wb = build_single_workbook(event)
        last = wb.active.max_row
        wb.save(dest_path)
        return TargetResult("new", dest_path, ok=True, last_row=last)
    except Exception as exc:  # noqa: BLE001 - 回報給 UI
        return TargetResult("new", dest_path, ok=False, error=str(exc))


def write_to_master(event: Event, master_path: str) -> TargetResult:
    name = _safe_sheet_title(event.sheet_name)
    try:
        if os.path.exists(master_path):
            wb = load_workbook(master_path)
        else:
            wb = Workbook()
            wb.remove(wb.active)

        replaced = name in wb.sheetnames
        if replaced:
            del wb[name]
        ws = wb.create_sheet(title=name)
        last = xlsx_writer.write_sheet(ws, event)
        wb.save(master_path)
        return TargetResult("master", master_path, ok=True,
                            replaced_sheet=replaced, last_row=last)
    except PermissionError:
        return TargetResult("master", master_path, ok=False,
                            error="無法寫入總表：檔案可能正在 Excel 中開啟，請關閉後再試。")
    except Exception as exc:  # noqa: BLE001
        return TargetResult("master", master_path, ok=False, error=str(exc))


def run(event: Event, *, to_master: bool = False, master_path: str = "",
        save_new: bool = False, dest_path: str = "") -> list[TargetResult]:
    """依勾選的目標一次輸出。至少要有一個目標。"""
    if not (to_master or save_new):
        raise ValueError("請至少選擇一個輸出目標（寫入總表／另存新檔）。")

    results: list[TargetResult] = []
    if to_master:
        if not master_path:
            results.append(TargetResult("master", "", ok=False, error="未指定總表路徑。"))
        else:
            results.append(write_to_master(event, master_path))
    if save_new:
        if not dest_path:
            results.append(TargetResult("new", "", ok=False, error="未指定另存路徑。"))
        else:
            results.append(save_as_new(event, dest_path))
    return results


def _safe_sheet_title(name: str) -> str:
    """Excel 工作表名稱限制：<=31 字元，且不可含 : \\ / ? * [ ]。"""
    bad = set(r':\/?*[]')
    cleaned = "".join("_" if ch in bad else ch for ch in name)
    return cleaned[:31] or "工作表"
