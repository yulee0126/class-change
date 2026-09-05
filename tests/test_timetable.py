import datetime
import os

import pytest

from tiaoke import storage, timetable as tt
from tiaoke.timetable import Slot, TeacherTable, Timetable, _klass_matches, _parse_cell
from tiaoke.ui.controller import AppController

D = datetime.date

# 這台機器上的實體 PDF（若不在則跳過相關測試）
_PDF = r"C:\Users\lolola\Desktop\1151教師課表_正式公布.pdf"
_HAS_PDF = os.path.exists(_PDF)

# 這台機器上的實體教師配當表（若不在則跳過相關測試）
_PEIDANG = r"C:\Users\lolola\Desktop\115-1教師配當表簽稿.xlsx"
_HAS_PEIDANG = os.path.exists(_PEIDANG)

# 從真實 PDF 掃出來的班級／地點詞彙（測 _parse_cell）
_CLASSES = {"園藝一", "園藝二", "園藝三", "高一甲", "高二甲", "商經一",
            "加工一", "餐飲一", "畜保一", "電機一", "綜職二"}
_ROOMS = {"園藝視聽教室", "多媒體教室2", "果樹技術教室", "室內配線場"}


def test_parse_cell_basic():
    r = _parse_cell(["農業概論", "園藝一", "園藝視聽教室"], _CLASSES, _ROOMS)
    assert r["subject"] == "農業概論"
    assert r["classes"] == ["園藝一"]
    assert r["location"] == "園藝視聽教室"


def test_parse_cell_wrapped_subject():
    r = _parse_cell(["農業資訊管理", "實習", "園藝三", "多媒體教室"], _CLASSES | {"園藝三"}, set())
    assert r["subject"] == "農業資訊管理實習"
    assert r["classes"] == ["園藝三"]


def test_parse_cell_meeting_skipped():
    assert _parse_cell(["行政會報"], _CLASSES, _ROOMS) is None


def test_parse_cell_multi_class_slash_split_across_lines():
    r = _parse_cell(["臺灣手語", "加工一/商", "經一/餐飲一/"], _CLASSES, set())
    assert r["subject"] == "臺灣手語"
    assert r["classes"] == ["加工一", "商經一", "餐飲一"]


def test_parse_cell_extra_hours_mark_on_class():
    r = _parse_cell(["農業概論", "園藝一(兼)", "園藝視聽教室"], _CLASSES, _ROOMS)
    assert r["subject"] == "農業概論"
    assert r["classes"] == ["園藝一"]
    assert r["note"] == "(兼)"


def test_parse_cell_extra_hours_mark_on_subject():
    r = _parse_cell(["數位科技", "概論(兼)", "高二丁應"], _CLASSES | {"高二丁應"}, set())
    assert r["subject"] == "數位科技概論"
    assert r["classes"] == ["高二丁應"]
    assert r["note"] == "(兼)"


def test_parse_cell_extra_hours_mark_own_line():
    r = _parse_cell(["應用數學", "(兼)", "畜保三"], _CLASSES | {"畜保三"}, set())
    assert r["subject"] == "應用數學"
    assert r["classes"] == ["畜保三"]
    assert r["note"] == "(兼)"


def test_parse_cell_tutoring_mark():
    r = _parse_cell(["數學輔導", "課 (輔)", "高二甲"], _CLASSES | {"高二甲"}, set())
    assert r["subject"] == "數學輔導課"
    assert r["classes"] == ["高二甲"]
    assert r["note"] == "(輔)"


def test_parse_cell_class_in_middle():
    r = _parse_cell(["電工實習", "電機一", "室內配線場"], _CLASSES, set())
    assert r["subject"] == "電工實習"
    assert r["classes"] == ["電機一"]
    assert r["location"] == "室內配線場"


def test_timetable_round_trip():
    src = Timetable(school="X中", semester="115學年度第一",
                    valid_from="2026-08-31", valid_to="2026-09-04",
                    period_times={1: ["08:00", "08:50"]})
    src.teachers["余瑞文"] = TeacherTable(
        name="余瑞文", tid="01306", title="兼課教師",
        slots=[Slot(1, 1, "健康與護理", "電機一", "", ""),
               Slot(3, 5, "健康與護理", "園藝一")])
    d = src.to_dict()
    back = Timetable.from_dict(d)
    assert back.teachers["余瑞文"].tid == "01306"
    assert back.teachers["余瑞文"].title == "兼課教師"
    assert back.slots_for("余瑞文", 1)[0].subject == "健康與護理"
    assert back.slots_for("余瑞文", 3)[0].klass == "園藝一"


def test_slot_group_reads_and_overwrites_combined_class():
    t = TeacherTable(name="王", slots=[
        Slot(1, 3, "客語", "加工一"), Slot(1, 3, "客語", "商經一"),
        Slot(1, 4, "國文", "高一甲"),
    ])
    assert {s.klass for s in t.slot_group(1, 3)} == {"加工一", "商經一"}
    assert len(t.slot_group(1, 4)) == 1

    t.set_slot_group(1, 3, "客語(四縣腔)", ["加工一", "餐飲一"], "語言教室", "(兼)")
    grp = t.slot_group(1, 3)
    assert {s.klass for s in grp} == {"加工一", "餐飲一"}
    assert all(s.subject == "客語(四縣腔)" and s.location == "語言教室" and s.note == "(兼)"
              for s in grp)
    # 沒動到的節次不受影響
    assert len(t.slot_group(1, 4)) == 1


def test_set_slot_group_single_class():
    t = TeacherTable(name="王")
    t.set_slot_group(2, 5, "數學", ["高二甲"])
    assert t.slots == [Slot(2, 5, "數學", "高二甲", "", "")]


def test_set_slot_group_with_co_teachers():
    t = TeacherTable(name="趙瑋")
    t.set_slot_group(2, 5, "基礎雜糧加工實作", ["綜職二"], co_teachers=["周蓁妍"])
    assert t.slots[0].co_teachers == ["周蓁妍"]

    # 改寫同一格、不再給 co_teachers → 清空
    t.set_slot_group(2, 5, "基礎雜糧加工實作", ["綜職二"])
    assert t.slots[0].co_teachers == []


def test_delete_slot_group():
    t = TeacherTable(name="王", slots=[
        Slot(1, 3, "客語", "加工一"), Slot(1, 3, "客語", "商經一"),
        Slot(1, 4, "國文", "高一甲"),
    ])
    t.delete_slot_group(1, 3)
    assert t.slot_group(1, 3) == []
    assert len(t.slots) == 1


def test_controller_edit_and_delete_timetable_slot():
    c = AppController()
    c.timetable = Timetable()
    c.timetable.teachers["王"] = TeacherTable(name="王", slots=[Slot(1, 2, "國文", "高一甲")])

    assert c.timetable_teacher_table("王").slots == [Slot(1, 2, "國文", "高一甲", "", "")]
    assert c.timetable_teacher_table("不存在") is None
    assert AppController().timetable_teacher_table("王") is None

    c.edit_timetable_slot("王", 1, 2, subject="國文", klasses=["高一甲", "高一乙"],
                          location="", note="(兼)", co_teachers=["李老師"])
    grp = c.timetable_teacher_table("王").slot_group(1, 2)
    assert {s.klass for s in grp} == {"高一甲", "高一乙"}
    assert all(s.note == "(兼)" and s.co_teachers == ["李老師"] for s in grp)

    c.delete_timetable_slot("王", 1, 2)
    assert c.timetable_teacher_table("王").slot_group(1, 2) == []


def test_controller_edit_timetable_slot_creates_teacher_and_timetable():
    c = AppController()
    assert c.timetable is None
    c.edit_timetable_slot("新老師", 2, 1, subject="化學", klasses=["高三甲"])
    assert c.timetable is not None
    assert c.timetable_teacher_table("新老師").slot_group(2, 1)[0].subject == "化學"


def test_controller_parse_co_teaching_requires_timetable_first(tmp_path):
    c = AppController()
    path = tmp_path / "配當表.xlsx"
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["header"])
    wb.save(path)
    with pytest.raises(ValueError, match="先匯入教師課表"):
        c.parse_co_teaching(str(path))


def test_controller_parse_co_teaching_applies_to_timetable(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["115-1教師配課一覽表"])
    ws.append(["序號", "職稱", "教師姓名", "授課班級", "課程名稱", "課程科別",
              "學分數", "基本節數", "兼課節數", "授課總數"])
    ws.append([1, "老師", "趙瑋", "職二", "基礎雜糧加工實作", 3, "協同", 12, 3, 15])
    ws.append([2, "老師", "周蓁妍", "職二", "基礎雜糧加工實作", 3, "協同", 8, 4, 12])
    path = tmp_path / "配當表.xlsx"
    wb.save(path)

    c = AppController()
    c.timetable = _co_teach_timetable()
    touched = c.parse_co_teaching(str(path))
    assert touched == 4
    assert c.timetable.teachers["趙瑋"].slots[0].co_teachers == ["周蓁妍"]


def test_controller_save_timetable_to(tmp_path):
    c = AppController()
    assert c.save_timetable_to(str(tmp_path / "x.json")) is None   # 沒課表
    c.timetable = Timetable()
    c.timetable.teachers["王"] = TeacherTable(name="王", slots=[Slot(1, 1, "國文", "高一甲")])
    path = c.save_timetable_to(str(tmp_path / "課表.json"))
    assert path and os.path.exists(path)
    back = storage.load_timetable(path)
    assert back.teachers["王"].slots[0].subject == "國文"


def test_klass_matches_short_code_against_full_name():
    assert _klass_matches("職二", "綜職二")
    assert _klass_matches("園一", "園藝一")
    assert _klass_matches("畜二", "畜保二")
    assert _klass_matches("工二", "加工二")
    assert not _klass_matches("園二", "高二")     # 字不對
    assert not _klass_matches("二一", "一二")     # 順序不對


def _co_teach_timetable() -> Timetable:
    table = Timetable()
    table.teachers["趙瑋"] = TeacherTable(name="趙瑋", slots=[
        Slot(2, 5, "基礎雜糧加工實作", "綜職二"),
        Slot(2, 6, "基礎雜糧加工實作", "綜職二"),
    ])
    table.teachers["周蓁妍"] = TeacherTable(name="周蓁妍", slots=[
        Slot(2, 5, "基礎雜糧加工實作", "綜職二"),
        Slot(2, 6, "基礎雜糧加工實作", "綜職二"),
    ])
    return table


def test_apply_co_teaching_links_matching_slot():
    table = _co_teach_timetable()
    rows = [
        ("趙瑋", "職二", "基礎雜糧加工實作", "協同"),
        ("周蓁妍", "職二", "基礎雜糧加工實作", "協同"),
    ]
    touched = tt.apply_co_teaching(table, rows)
    assert touched == 4  # 兩位老師各 2 節
    for name in ("趙瑋", "周蓁妍"):
        other = "周蓁妍" if name == "趙瑋" else "趙瑋"
        for s in table.teachers[name].slots:
            assert s.co_teachers == [other]


def test_apply_co_teaching_confirms_even_if_only_one_side_flagged():
    """配當表常常只有一邊標協同，另一邊完全沒標記——只要同節次對得上就該算數。"""
    table = _co_teach_timetable()
    rows = [
        ("趙瑋", "職二", "基礎雜糧加工實作", "協同"),
        ("周蓁妍", "職二", "基礎雜糧加工實作", ""),  # 沒標記
    ]
    touched = tt.apply_co_teaching(table, rows)
    assert touched == 4


def test_apply_co_teaching_rejects_without_any_co_teach_flag():
    """兩位老師同班同課，但配當表完全沒標協同（例如各自帶不同組的專題）→ 不算協同。"""
    table = _co_teach_timetable()
    rows = [
        ("趙瑋", "職二", "基礎雜糧加工實作", ""),
        ("周蓁妍", "職二", "基礎雜糧加工實作", ""),
    ]
    touched = tt.apply_co_teaching(table, rows)
    assert touched == 0
    for name in ("趙瑋", "周蓁妍"):
        assert all(s.co_teachers == [] for s in table.teachers[name].slots)


def test_apply_co_teaching_rejects_when_periods_dont_actually_match():
    """同班同課、有標協同，但實際課表節次對不上（例如各自不同時段帶開）→ 不算協同。"""
    table = Timetable()
    table.teachers["甲"] = TeacherTable(name="甲", slots=[Slot(1, 1, "專題實作", "畜二")])
    table.teachers["乙"] = TeacherTable(name="乙", slots=[Slot(3, 4, "專題實作", "畜二")])
    rows = [
        ("甲", "畜二", "專題實作", "協同"),
        ("乙", "畜二", "專題實作", "協同"),
    ]
    touched = tt.apply_co_teaching(table, rows)
    assert touched == 0


def test_parse_co_teaching_xlsx(tmp_path):
    """欄位順序照真實檔案：G 欄（index 6）是協同/分組標記，不是「課程科別」。"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["115-1教師配課一覽表"])
    ws.append(["序號", "職稱", "教師姓名", "授課班級", "課程名稱", "課程科別",
              "學分數", "基本節數", "兼課節數", "授課總數"])
    ws.append([1, "老師", "趙瑋", "職二", "基礎雜糧加工實作", 3, "協同", 12, 3, 15])
    ws.append([None, None, None, "職三", "雜糧作物加工實習", 3, None, None, None, None])
    ws.append([2, "老師", "周蓁妍", "職二", "基礎雜糧加工實作", 3, "協同", 8, 4, 12])
    path = tmp_path / "配當表.xlsx"
    wb.save(path)

    rows = tt.parse_co_teaching_xlsx(str(path))
    assert ("趙瑋", "職二", "基礎雜糧加工實作", "協同") in rows
    assert ("趙瑋", "職三", "雜糧作物加工實習", "") in rows
    assert ("周蓁妍", "職二", "基礎雜糧加工實作", "協同") in rows


def test_parse_co_teaching_xlsx_picks_sheet_with_most_markers(tmp_path):
    """同一活頁簿裡若有舊學期分頁（表頭一樣但沒有協同標記），要挑對現在在用的那個。"""
    import openpyxl
    wb = openpyxl.Workbook()
    old = wb.active
    old.title = "工作表1"
    header = ["序號", "職稱", "教師姓名", "授課班級", "課程名稱", "課程科別", "學分數",
             "基本節數", "兼課節數", "授課總數"]
    old.append(["114-1教師配課一覽表"])
    old.append(header)
    old.append([1, "老師", "藍秋月", "園一", "植物栽培實習", "園藝", 3, 8, 4, 11])

    new = wb.create_sheet("工作表2")
    new.append(["115-1教師配課一覽表"])
    new.append(header)
    new.append([1, "老師", "趙瑋", "職二", "基礎雜糧加工實作", 3, "協同", 12, 3, 15])
    new.append([2, "老師", "周蓁妍", "職二", "基礎雜糧加工實作", 3, "協同", 8, 4, 12])

    path = tmp_path / "配當表.xlsx"
    wb.save(path)

    rows = tt.parse_co_teaching_xlsx(str(path))
    assert ("趙瑋", "職二", "基礎雜糧加工實作", "協同") in rows
    assert not any(r[0] == "藍秋月" for r in rows)


@pytest.mark.skipif(not (_HAS_PEIDANG and _HAS_PDF), reason="找不到實體配當表或課表 PDF")
def test_apply_co_teaching_real_files():
    table = tt.parse_pdf(_PDF)
    rows = tt.parse_co_teaching_xlsx(_PEIDANG)
    touched = tt.apply_co_teaching(table, rows)
    assert touched > 0

    zhao = table.teachers.get("趙瑋")
    zhen = table.teachers.get("周蓁妍")
    assert zhao and zhen
    linked = [s for s in zhao.slots if s.subject == "基礎雜糧加工實作" and s.co_teachers]
    assert linked and all(s.co_teachers == ["周蓁妍"] for s in linked)
    linked_back = [s for s in zhen.slots if s.subject == "基礎雜糧加工實作" and s.co_teachers]
    assert linked_back and all(s.co_teachers == ["趙瑋"] for s in linked_back)

    # 已知的「同課程但各自不同節次」假陽性案例不該被標記
    for name in table.teachers:
        for s in table.teachers[name].slots:
            if s.subject == "專題初探" and "畜" in s.klass:
                assert s.co_teachers == []


def test_controller_timetable_slots_weekday_mapping():
    c = AppController()
    c.timetable = Timetable()
    c.timetable.teachers["王"] = TeacherTable(name="王", slots=[
        Slot(1, 2, "國文", "高一甲"),   # 星期一
        Slot(3, 4, "國文", "高一乙"),   # 星期三
    ])
    # 2026-08-31 是星期一
    got = c.timetable_slots("王", D(2026, 8, 31))
    assert [s.period for s in got] == [2]
    # 星期六 → 無
    assert c.timetable_slots("王", D(2026, 9, 5)) == []
    # 沒課表
    assert AppController().timetable_slots("王", D(2026, 8, 31)) == []


@pytest.mark.skipif(not _HAS_PDF, reason="找不到實體課表 PDF")
def test_parse_real_pdf():
    table = tt.parse_pdf(_PDF)
    assert len(table.teachers) >= 80
    assert table.school == "國立關西高級中學"
    assert table.valid_from == "2026-09-07"
    yu = table.teachers.get("余瑞文")
    assert yu and all(s.subject == "健康與護理" for s in yu.slots)
    # 余瑞文全部課都是兼課（頁尾：基本鐘點0堂／兼課8堂）
    assert all(s.note == "(兼)" for s in yu.slots)
    # 星期一應有 4 節
    assert len(yu.on(1)) == 4
    # 空班級的比例應該很低（這份 PDF 有少數幾格因跨行排版把 (兼) 黏進文字中間而解析失敗，
    # 修正前的舊邏輯是 11.9%，修正後約 2.3%）
    empties = [s for t in table.teachers.values() for s in t.slots if not s.klass]
    total = sum(len(t.slots) for t in table.teachers.values())
    assert len(empties) / total < 0.03
    # (兼)/(輔) 標記不應殘留在科目或班級文字裡（應已搬進 note）
    for t in table.teachers.values():
        for s in t.slots:
            assert "(兼)" not in s.subject and "(兼)" not in s.klass
            assert "(輔)" not in s.subject and "(輔)" not in s.klass
    # 校內慣例：第八節一律算輔導，不論 PDF 該格是否印出 (輔)
    for t in table.teachers.values():
        for s in t.slots:
            if s.period == 8:
                assert "(輔)" in s.note
